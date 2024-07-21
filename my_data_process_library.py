from urllib.error import URLError
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
from tabulate import tabulate
import urllib
import datetime
def convert_date(date_str):
    # print(date_str)
    if isinstance(date_str, datetime.date):
        # If the input is a date object, swap month and day
        return datetime.date(date_str.year, date_str.day, date_str.month)
    elif isinstance(date_str, str):
        # If the input is a string, extract day, month, and year
        
        try:
            day, month, year = int(date_str[:2]), int(date_str[3:5]), int(date_str[6:10])
            return datetime.date(year, month, day)
        except ValueError:
            return datetime.date(1999, 9, 9)
    else:
        return datetime.date(1999, 9, 9)

def get_UN_data(LINK_PUBLIC_TO_WEB: str):
    # Hàm load file online, khi published
    try:
        # AWS_BUCKET_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQNpA9xv7ci1tGPdF1I-HwPdPWNvyryr5YNQvXOwxKRIWdOg5zPy-2xvXjrRoChqeb6QmwQX-qO4-uO/pub?output=csv"
        # LINK_PUBLIC_TO_WEB = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQNpA9xv7ci1tGPdF1I-HwPdPWNvyryr5YNQvXOwxKRIWdOg5zPy-2xvXjrRoChqeb6QmwQX-qO4-uO/pub?output=xlsx"
        # AWS_BUCKET_URL = "https://github.com/drquochoai/Streamlit-thong-ke-phau-thuat/raw/main/Thang%206%20PTTT.xlsx"
        LINK_EDIT = "https://docs.google.com/spreadsheets/d/18k646SexxQPgnhO6R4s1XghRitHrp2yFqynWSK9GqK8/edit"
        with urllib.request.urlopen(LINK_PUBLIC_TO_WEB) as response:
            file_object_load_GG_SHEET_directed_to_variable = io.BytesIO(
                response.read())
        # with open("temp_file.xlsx", "wb") as f:
        #     f.write(html)

        # Load the Excel file using openpyxl
        wb = load_workbook(file_object_load_GG_SHEET_directed_to_variable)
        # List all sheet names
        sheet_names = wb.sheetnames
        print("Sheet names:")
        for sheet_name in sheet_names:
            print(f"- {sheet_name}")
        option_Sheet_thong_ke = st.selectbox(
            "Chọn sheet muốn thống kê:",
            wb.sheetnames)

        df = pd.read_excel(
            file_object_load_GG_SHEET_directed_to_variable, sheet_name=option_Sheet_thong_ke, engine='openpyxl')
        # inf_bao = st.warning("Dữ liệu đã được tải thành công.")
        inf_moLinkEdit = st.link_button("Mở trang dữ liệu", url=LINK_EDIT)
        # df['NGAY'] = df['NGAY'].astype(str)
        if 'SONHA' in df.columns:
            df['SONHA'] = df['SONHA'].astype(str)
        if 'SOCMND' in df.columns:
            df['SOCMND'] = df['SOCMND'].astype(str)
        if 'NGAYCAP' in df.columns:
            df['NGAYCAP'] = df['NGAYCAP'].astype(str)
        if 'NGAY' in df.columns:
            df['NGAY'] = df['NGAY'].apply(convert_date)
        if 'NGAYKT' in df.columns:
            df['NGAYKT'] = df['NGAYKT'].apply(convert_date)
        if 'NGAYRV' in df.columns:
            df['NGAYRV'] = df['NGAYRV'].apply(convert_date)
        if 'NGAYRUT' in df.columns:
            df['NGAYRUT'] = df['NGAYRUT'].apply(convert_date)
        if 'NGAYCATCHI' in df.columns:
            df['NGAYCATCHI'] = df['NGAYCATCHI'].apply(convert_date)
        return df.reset_index(drop=True)

        with urllib.request.urlopen(LINK_PUBLIC_TO_WEB) as f:
            html = f.read()
    except URLError as e:
        st.error(f"Lỗi load file từ google sheet: {e}")

    # Hàm load file từ local, khi debug

    uploaded_file = st.file_uploader("Tải file xlSX")
    if uploaded_file is not None:
        df = pd.read_excel(uploaded_file, sheet_name=0, engine='openpyxl')
        inf_bao.empty()
        return df.set_index("ID")
